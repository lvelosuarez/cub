import marimo

__generated_with = "0.19.6"
app = marimo.App(width="full")


@app.cell
def _():
    from pathlib import Path

    import marimo as mo
    import polars as pl
    from openpyxl import Workbook

    return Path, Workbook, mo, pl


@app.cell(hide_code=True)
def _(Path, mo):
    file_browser = mo.ui.file_browser(
        initial_path=Path("/mnt/san/microbio/m_clinique/"),  # where the browser opens
        selection_mode="file",  # or "directory"
        multiple=True,  # allow selecting several files
        filetypes=[".tsv", ".csv"],  # optional filter, or None for all
        label="Select your input files",
    )
    file_browser  # last expression -> shown in UI
    return (file_browser,)


@app.cell(hide_code=True)
def _(file_browser, mo):
    # If nothing is selected yet
    if not file_browser.value:
        selected_paths = []
    else:
        # file_browser.value is a list-like of "file info" objects
        # We turn it into a list of real paths
        selected_paths = [
            file_browser.path(i)  # absolute path for item i
            for i, _info in enumerate(file_browser.value)
        ]

    # Optional: show them nicely in the UI
    if not selected_paths:
        display = mo.md("No files selected yet.")
    else:
        lines = "\n".join(f"- `{p}`" for p in selected_paths)
        display = mo.md("### Selected files\n" + lines)

    display  # rendered in UI
    return (selected_paths,)


@app.cell
def _(mo):
    mo.md(r"""
    # Count reads
    """)
    return


@app.cell
def _(selected_paths):
    selected_paths[2]
    return


@app.cell
def _(pl, selected_paths):
    df_r = pl.read_csv(selected_paths[2], separator=",")
    df_r
    return (df_r,)


@app.cell
def _(pl, selected_paths):
    order = ["illumina", "qc", "dedup", "dehost", "merge"]
    df = pl.read_csv(selected_paths[2], separator="\t")
    df = (
        df.with_columns(sample=pl.col("file").str.extract(r"(MG-Run\d+-Ech-\d+)", 1))
        .with_columns(
            step=pl.when(pl.col("file").str.contains("output/merge"))
            .then(pl.lit("merge"))
            .when(pl.col("file").str.contains("output/qc"))
            .then(pl.lit("qc"))
            .when(pl.col("file").str.contains("output/dedup"))
            .then(pl.lit("dedup"))
            .when(pl.col("file").str.contains("output/dehost"))
            .then(pl.lit("dehost"))
            .otherwise(pl.lit("illumina"))
        )
        .group_by(["sample", "step"])
        .agg(pl.col("output").sum().alias("output"))
        .group_by(["sample", "step"])
        .agg(pl.col("output").max().alias("reads"))
        .pivot(
            values="reads",
            index="sample",
            on="step",
        )
        .select(["sample"] + order)
        .with_columns([pl.col(c).fill_null(0).cast(pl.Int64) for c in order])
        .sort("sample")
        .with_columns(pl.col("sample").str.extract(r"(\d+)$").cast(pl.Int32).alias("id"))
        .sort(by=["id"])
        .drop("id")
        .filter(pl.col("sample").is_not_null())
    )
    return df, order


@app.cell(hide_code=True)
def _(df, order, pl):
    (
        df.with_columns(sample=pl.col("file").str.extract(r"(MG-Run\d+-Ech-\d+)", 1))
        .with_columns(
            step=pl.when(pl.col("file").str.contains("output/merge"))
            .then(pl.lit("merge"))
            .when(pl.col("file").str.contains("output/qc"))
            .then(pl.lit("qc"))
            .when(pl.col("file").str.contains("output/dedup"))
            .then(pl.lit("dedup"))
            .when(pl.col("file").str.contains("output/dehost"))
            .then(pl.lit("dehost"))
            .otherwise(pl.lit("illumina"))
        )
        .group_by(["sample", "step"])
        .agg(pl.col("output").sum().alias("output"))
        .group_by(["sample", "step"])
        .agg(pl.col("output").max().alias("reads"))
        .pivot(
            values="reads",
            index="sample",
            on="step",
        )
        .select(["sample"] + order)
        .with_columns([pl.col(c).fill_null(0).cast(pl.Int64) for c in order])
        .sort("sample")
        .with_columns(pl.col("sample").str.extract(r"(\d+)$").cast(pl.Int32).alias("id"))
        .sort(by=["id"])
        .drop("id")
        .filter(pl.col("sample").is_not_null())
    )
    return


@app.cell
def _(df, order, pl):
    _df = (
        df.with_columns(pl.col("file").str.split("/").list.get(-1).alias("basename"))
        .with_columns(pl.col("basename").str.replace(r"_dedup.*$", "").str.replace(r"_R1.*$", "").alias("sample"))
        .with_columns(
            pl.when(pl.col("file").str.starts_with("output/qc/"))
            .then(pl.lit("qc"))
            .when(pl.col("file").str.starts_with("output/merge/"))
            .then(pl.lit("merge"))
            .when(pl.col("file").str.starts_with("output/dedup/"))
            .then(pl.lit("dedup"))
            .when(pl.col("file").str.starts_with("output/dehost/"))
            .then(pl.lit("dehost"))
            .otherwise(pl.lit("illumina"))
            .alias("stage")
        )
        .group_by(["sample", "stage"])
        .agg(pl.col("output").max().alias("reads"))
        .pivot(
            values="reads",
            index="sample",
            on="stage",
        )
        .select(["sample"] + order)
        .with_columns([pl.col(c).fill_null(0).cast(pl.Int64) for c in order])
        .sort("sample")
        .with_columns(pl.col("sample").str.extract(r"(\d+)_").cast(pl.Int32).alias("id"))
        .sort(by=["id"])
        .drop("id")
    )
    return


@app.cell
def _(df, pl):
    df.with_columns(pl.col("sample").str.extract(r"(\d+)_").cast(pl.Int32).alias("id")).sort(by=["id"]).drop("id")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # Analysis
    """)
    return


@app.function
def write_polars_df(ws, df):
    # write header
    ws.append(df.columns)

    # write rows
    for row in df.iter_rows():
        ws.append(row)


@app.cell
def _(selected_paths):
    # selected_paths[0]
    selected_paths[1]
    # selected_paths[2]
    return


@app.cell
def _(selected_paths):
    KRAKEN1 = selected_paths[1]
    # KRAKEN2 = selected_paths[2]
    BAM = selected_paths[0]
    has_neg = False
    return BAM, KRAKEN1, has_neg


@app.cell
def _(KRAKEN1, pl):
    kraken = pl.read_csv(KRAKEN1, separator="\t")
    # kraken2 = pl.read_csv(KRAKEN2, separator="\t")
    # kraken = pl.concat([kraken1, kraken2])
    kraken_sel = (
        kraken.filter(pl.col("fragments_direct") != 0)
        .select("sample", "taxid", "name", "fragments_direct", "genome_coverage_estimate", "z_score", "p_value", "n_samples_with_taxon", "rank_code")
        .with_columns(pl.col("taxid").cast(pl.Int64, strict=False))
        .rename(
            {
                "z_score": "kraken_z",
                "p_value": "kraken_p",
            }
        )
    )
    return (kraken_sel,)


@app.cell
def _(BAM, pl):
    bam = pl.read_csv(BAM, separator="\t")
    bam_sel = (
        bam.select("sample", "ncbi_taxid", "gtdb_taxonomy", "read_count", "z_score", "p_value")
        .with_columns(pl.col("ncbi_taxid").cast(pl.Int64, strict=False).alias("taxid"))
        .with_columns(pl.when(pl.col("taxid") == 765068).then(1747).otherwise(pl.col("taxid")).alias("taxid"))
        .rename(
            {
                "z_score": "bam_z",
                "p_value": "bam_p",
            }
        )
        .drop("ncbi_taxid")
    )
    return (bam_sel,)


@app.cell
def _(bam_sel, has_neg, kraken_sel, pl):
    merged = (
        bam_sel.join(kraken_sel, on=["sample", "taxid"], how="full")
        .with_columns(
            # Ensure keys are always present
            pl.coalesce([pl.col("sample"), pl.col("sample_right")]).alias("sample"),
            pl.coalesce([pl.col("taxid"), pl.col("taxid_right")]).alias("taxid"),
        )
        .select(
            "sample",
            "taxid",
            "gtdb_taxonomy",
            "name",
            "bam_z",
            "bam_p",
            "read_count",
            "fragments_direct",
            "kraken_z",
            "kraken_p",
            "n_samples_with_taxon",
            "genome_coverage_estimate",
        )
    )

    # Mask statistics when there is NO negative sample
    if not has_neg:
        merged = merged.with_columns(
            pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("bam_z")).alias("bam_z"),
            pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("bam_p")).alias("bam_p"),
            pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("kraken_z")).alias("kraken_z"),
            pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("kraken_p")).alias("kraken_p"),
        )

    merged = (
        merged.with_columns(
            pl.col("bam_z").round(2),
            pl.col("bam_p").round(2),
            pl.col("kraken_p").round(2),
            pl.col("kraken_z").round(2),
            pl.col("genome_coverage_estimate").round(6),
        )
        .rename(
            {
                "read_count": "bam_reads",
                "fragments_direct": "kraken_reads",
            }
        )
        # keep non-host AND rows where name is missing
        .filter(pl.col("name").is_null() | (pl.col("name") != "Homo sapiens"))
        .sort(by=["sample", "bam_reads", "kraken_reads"], descending=[False, True, True])
    )
    return (merged,)


@app.cell
def _(merged, pl):
    N = 30  # max hits per sample – adjust as you like

    best_hits = (
        merged
        # 1) fill null reads
        .with_columns(
            pl.col("bam_reads").fill_null(0).alias("bam_reads"),
            pl.col("kraken_reads").fill_null(0).alias("kraken_reads"),
        )
        # 2) flags
        .with_columns(
            (pl.col("bam_reads") > 0).alias("has_bam"),
            (pl.col("kraken_reads") > 0).alias("has_kraken"),
            # Kraken stats exist if both z and p are non-null
            (pl.col("kraken_z").is_not_null() & pl.col("kraken_p").is_not_null()).alias("has_kraken_stats"),
        )
        # 3) evidence category + combined score (for ranking only)
        .with_columns(
            # 2 = both, 1 = kraken-only, 0 = bam-only
            pl.when(pl.col("has_bam") & pl.col("has_kraken"))
            .then(2)
            .when(~pl.col("has_bam") & pl.col("has_kraken"))
            .then(1)
            .otherwise(0)
            .alias("evidence_cat"),
            pl.when(pl.col("has_bam") & pl.col("has_kraken"))
            .then(pl.col("bam_z").fill_null(0) + pl.col("kraken_z").fill_null(0))
            .when(pl.col("has_bam"))
            .then(pl.col("bam_z").fill_null(0))
            .otherwise(pl.col("kraken_z").fill_null(0))
            .alias("combined_score"),
        )
        # 4) FILTER according to your truth table
        .filter(
            # 1) no BAM / has Kraken / has stats -> use Kraken stats
            (~pl.col("has_bam") & pl.col("has_kraken") & pl.col("has_kraken_stats") & (pl.col("kraken_z") >= 1) & (pl.col("kraken_p") <= 0.1))
            |
            # 2) no BAM / has Kraken / no stats -> filter Kraken reads
            (
                ~pl.col("has_bam") & pl.col("has_kraken") & ~pl.col("has_kraken_stats") & (pl.col("kraken_reads") >= 10)  # threshold à régler
            )
            |
            # 3) has BAM / has Kraken / has stats -> use stats (both)
            (
                pl.col("has_bam")
                & pl.col("has_kraken")
                & pl.col("has_kraken_stats")
                # & (pl.col("bam_z") >= 1)
                # & (pl.col("bam_p") <= 0.1)
                & (pl.col("kraken_z") >= 1.5)
                & (pl.col("kraken_p") <= 0.05)
            )
            |
            # 4) has BAM / has Kraken / no stats -> keep ALL
            (pl.col("has_bam") & pl.col("has_kraken") & ~pl.col("has_kraken_stats"))
            |
            # 5) has BAM / no Kraken -> use BAM stats
            (pl.col("has_bam") & ~pl.col("has_kraken") & (pl.col("bam_z") >= 1) & (pl.col("bam_p") <= 0.05))
        )
        # 5) rank + top N
        .sort(
            by=["sample", "evidence_cat", "combined_score"],
            descending=[False, True, True],
        )
        .group_by("sample", maintain_order=True)
        .head(N)
        .drop(["has_bam", "has_kraken", "has_kraken_stats", "evidence_cat", "combined_score"])
        .unique()
    )
    return (best_hits,)


@app.cell
def _(best_hits, pl):
    cols_order = [
        "sample",
        "name",
        "bam_z",
        "bam_p",
        "bam_reads",
        "kraken_reads",
        "kraken_z",
        "kraken_p",
        "n_samples_with_taxon",
        "genome_coverage_estimate",
        "taxid",
        "gtdb_taxonomy",
    ]
    final_df = (
        best_hits.with_columns(pl.col("gtdb_taxonomy").str.extract(r"s__([^;_]+(?: [^;_]+)?)", 1).alias("species"))
        .with_columns(
            pl.when(pl.col("name").is_not_null() & (pl.col("name") != ""))
            .then(pl.col("name"))
            .when(pl.col("species").is_not_null() & (pl.col("species") != ""))
            .then(pl.col("species"))
            .otherwise(None)
            .alias("name")
        )
        .drop("species")
        .unique()
        .with_columns(pl.col("sample").str.extract(r"(\d+)$").cast(pl.Int32).alias("id"))
        .sort(by=["id"])
        .select(cols_order)
    )
    return (final_df,)


@app.cell
def _(final_df):
    final_df
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # Export
    """)
    return


@app.cell
def _(Workbook, df_r, final_df, merged):
    wb = Workbook()

    # remove default sheet
    wb.remove(wb.active)

    # sheets
    ws_counts = wb.create_sheet("counts")
    ws_best = wb.create_sheet("selected_hits")
    ws_merged = wb.create_sheet("all_hits")

    # write data
    write_polars_df(ws_counts, df_r)
    write_polars_df(ws_best, final_df)
    write_polars_df(ws_merged, merged)

    # save
    wb.save("/mnt/san/microbio/m_clinique/NOSCENDO/260319_NB501647_0367_AHNYCHBGYX/alignment_lou/output/results.xlsx")
    return


if __name__ == "__main__":
    app.run()

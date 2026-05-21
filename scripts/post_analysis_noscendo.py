#!/usr/bin/env python3
"""
post_analysis_noscendo.py
Adapted from post_analysis.py for single-end, single-cub-file NOSCENDO run.
"""

import polars as pl
from pathlib import Path
from openpyxl import Workbook

# --- Paths ---
RUN_DIR = Path("/mnt/san/microbio/m_clinique/NOSCENDO/260305_NB501647_0364_AHNYWLBGYX")
BAM     = RUN_DIR / "Aligment_lou3/output/bam.tsv"
KRAKEN  = RUN_DIR / "cub.tsv"
NREADS  = RUN_DIR / "Aligment_lou3/output/Nreads.csv"
OUTPUT  = RUN_DIR / "results.xlsx"

has_neg = True
N = 30  # max hits per sample

# --- Load ---
df_r   = pl.read_csv(NREADS)
kraken = pl.read_csv(KRAKEN, separator="\t")
bam    = pl.read_csv(BAM,    separator="\t")

# --- Kraken selection ---
kraken_sel = (
    kraken.filter(pl.col("fragments_direct") != 0)
    .select(
        "sample", "taxid", "name", "fragments_direct",
        "genome_coverage_estimate", "z_score", "p_value",
        "n_samples_with_taxon", "rank_code",
    )
    .with_columns(pl.col("taxid").cast(pl.Int64, strict=False))
    .rename({"z_score": "kraken_z", "p_value": "kraken_p"})
)

# --- BAM selection ---
bam_sel = (
    bam.select("sample", "ncbi_taxid", "gtdb_taxonomy", "read_count", "z_score", "p_value")
    .with_columns(pl.col("ncbi_taxid").cast(pl.Int64, strict=False).alias("taxid"))
    .with_columns(
        pl.when(pl.col("taxid") == 765068).then(1747).otherwise(pl.col("taxid")).alias("taxid")
    )
    .rename({"z_score": "bam_z", "p_value": "bam_p"})
    .drop("ncbi_taxid")
)

# --- Merge ---
merged = (
    bam_sel.join(kraken_sel, on=["sample", "taxid"], how="full")
    .with_columns(
        pl.coalesce([pl.col("sample"), pl.col("sample_right")]).alias("sample"),
        pl.coalesce([pl.col("taxid"),  pl.col("taxid_right") ]).alias("taxid"),
    )
    .select(
        "sample", "taxid", "gtdb_taxonomy", "name",
        "bam_z", "bam_p", "read_count", "fragments_direct",
        "kraken_z", "kraken_p", "n_samples_with_taxon", "genome_coverage_estimate",
    )
)

if not has_neg:
    merged = merged.with_columns(
        pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("bam_z")).alias("bam_z"),
        pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("bam_p")).alias("bam_p"),
        pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("kraken_z")).alias("kraken_z"),
        pl.when(pl.col("n_samples_with_taxon") <= 1).then(None).otherwise(pl.col("kraken_p")).alias("kraken_p"),
    )

merged = (
    merged
    .with_columns(
        pl.col("bam_z").round(2),
        pl.col("bam_p").round(2),
        pl.col("kraken_p").round(2),
        pl.col("kraken_z").round(2),
        pl.col("genome_coverage_estimate").round(6),
    )
    .rename({"read_count": "bam_reads", "fragments_direct": "kraken_reads"})
    .filter(pl.col("name").is_null() | (pl.col("name") != "Homo sapiens"))
    .sort(by=["sample", "bam_reads", "kraken_reads"], descending=[False, True, True])
)

# --- Best hits ---
best_hits = (
    merged
    .with_columns(
        pl.col("bam_reads").fill_null(0).alias("bam_reads"),
        pl.col("kraken_reads").fill_null(0).alias("kraken_reads"),
    )
    .with_columns(
        (pl.col("bam_reads")    > 0).alias("has_bam"),
        (pl.col("kraken_reads") > 0).alias("has_kraken"),
        (pl.col("kraken_z").is_not_null() & pl.col("kraken_p").is_not_null()).alias("has_kraken_stats"),
    )
    .with_columns(
        pl.when(pl.col("has_bam") & pl.col("has_kraken")).then(2)
          .when(~pl.col("has_bam") & pl.col("has_kraken")).then(1)
          .otherwise(0).alias("evidence_cat"),
        pl.when(pl.col("has_bam") & pl.col("has_kraken"))
          .then(pl.col("bam_z").fill_null(0) + pl.col("kraken_z").fill_null(0))
          .when(pl.col("has_bam")).then(pl.col("bam_z").fill_null(0))
          .otherwise(pl.col("kraken_z").fill_null(0)).alias("combined_score"),
    )
    .filter(
        (~pl.col("has_bam") & pl.col("has_kraken") & pl.col("has_kraken_stats")
         & (pl.col("kraken_z") >= 1) & (pl.col("kraken_p") <= 0.1))
        | (~pl.col("has_bam") & pl.col("has_kraken") & ~pl.col("has_kraken_stats")
           & (pl.col("kraken_reads") >= 10))
        | (pl.col("has_bam") & pl.col("has_kraken") & pl.col("has_kraken_stats")
           & (pl.col("kraken_z") >= 1.5) & (pl.col("kraken_p") <= 0.05))
        | (pl.col("has_bam") & pl.col("has_kraken") & ~pl.col("has_kraken_stats"))
        | (pl.col("has_bam") & ~pl.col("has_kraken")
           & (pl.col("bam_z") >= 1) & (pl.col("bam_p") <= 0.05))
    )
    .sort(by=["sample", "evidence_cat", "combined_score"], descending=[False, True, True])
    .group_by("sample", maintain_order=True).head(N)
    .drop(["has_bam", "has_kraken", "has_kraken_stats", "evidence_cat", "combined_score"])
    .unique()
)

# --- Final formatting ---
cols_order = [
    "sample", "name", "bam_z", "bam_p", "bam_reads", "kraken_reads",
    "kraken_z", "kraken_p", "n_samples_with_taxon", "genome_coverage_estimate",
    "taxid", "gtdb_taxonomy",
]
final_df = (
    best_hits
    .with_columns(
        pl.col("gtdb_taxonomy").str.extract(r"s__([^;_]+(?: [^;_]+)?)", 1).alias("species")
    )
    .with_columns(
        pl.when(pl.col("name").is_not_null() & (pl.col("name") != ""))
          .then(pl.col("name"))
          .when(pl.col("species").is_not_null() & (pl.col("species") != ""))
          .then(pl.col("species"))
          .otherwise(None).alias("name")
    )
    .drop("species")
    .unique()
    .with_columns(pl.col("sample").str.extract(r"(\d+)$").cast(pl.Int32).alias("id"))
    .sort(by=["id"])
    .select(cols_order)
)

# --- Export ---
def write_polars_df(ws, df):
    ws.append(df.columns)
    for row in df.iter_rows():
        ws.append(row)

wb = Workbook()
wb.remove(wb.active)
write_polars_df(wb.create_sheet("counts"),        df_r)
write_polars_df(wb.create_sheet("selected_hits"), final_df)
write_polars_df(wb.create_sheet("all_hits"),      merged)
wb.save(str(OUTPUT))
print(f"Saved: {OUTPUT}")
